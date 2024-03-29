import pandas as pd
import numpy as np

import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.gridspec as gridspec

from openpyxl.styles import Alignment,Font,PatternFill
from openpyxl import Workbook, utils
from openpyxl.drawing.image import Image

from tkinter import filedialog

import re


class xlwriter:
    def __init__(self, df, views, names, suppressed_data):
        self.data = df
        self.views = views
        self.names = names
        self.complete_data = self.data.join(suppressed_data)
        self.output_row = 1
        self.img_idx = 0

                
    #Dictionary for translating day abbreviations
        self.days_dict = {'Monday': 'M', 'Tuesday': 'T', 
                    'Wednesday': 'W', 'Thursday': 'Th', 'Friday':'F'}
        self.days_dict_reverse = {v: k for k, v in self.days_dict.items()}

        self.dataframes = {}
        self.wb = Workbook()


    def auto_resize(self,ws):
    #Function to resize cells for legibility: double height and increase width by some 
    #percentage of the width of the longest contents in the column
        for row in ws:
            for cell in row:
                cell.alignment = Alignment(wrap_text=True,
                                        horizontal='left', 
                                        vertical='center'
                                        )
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length)/2.25 + 5
            try:
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            except AttributeError:
                pass

    
    def run(self):

        first = True

        for key, view in self.views.items():

        #Check if we're on the first sheet. If not, append
        #a new sheet and write to that
            if view[1].get() and first:
                self.ws = self.wb.active
                self.ws.title = view[0]

            elif view[1].get():
                self.ws = self.wb.create_sheet()
                self.ws.title = view[0]
            first = False

        #Write 'Course Schedule' sheet
            if key == 'default' and view[1].get(): self.write_course_schedule()

        #Write the 'By Instructor' sheet
            if key == 'by_instructor' and view[1].get(): self.write_by_instructor()

        #Write the 'By Day' sheet
            if key == 'by_day' and view[1].get(): self.write_by_day()

        #Write the 'Graphical Schedule' sheet
            
            if key == 'graphical' and view[1].get(): 
                self.write_graphical()

        self.save()


    def merge_headers(self,ws,column_values,end_col):
        for row in ws:
            if any(cell.value in column_values for cell in row):
                ws.merge_cells(start_column=1,
                               end_column=end_col,
                               start_row=row[1].row, end_row=row[1].row
                               )


    def save(self):
        filepath = filedialog.asksaveasfilename(initialfile='matrix.xlsx',defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        self.wb.save(filepath)

        
    def write_graphical(self):
        #Color palette for class bars
        palette = [
            '#FF6B6B',  # Light Coral
            '#87CEEB',  # Light Sky Blue
            '#90EE90',  # Light Green
            '#FFA07A',  # Light Salmon
            '#E6A8D7',  # Light Orchid
            '#FAFAD2',  # Light Goldenrod Yellow
            '#E0FFFF',  # Light Cyan
            '#FFB6C1',  # Light Pink
            '#778899',  # Light Slate Gray
            '#B0C4DE',  # Light Steel Blue
            '#FFFFE0',  # Light Yellow
            '#20B2AA',  # Light Sea Green
            '#E6E6FA'   # Light Lavender
        ]

        for term in self.complete_data['Term'].unique():
            data_term = self.complete_data[self.complete_data['Term'] == term]
            data_term['color'] = palette * (len(data_term) // len(palette)) + palette[:len(data_term) % len(palette)]
            #This is a bit stupid, but I run the same code twice to determine
            #how many columns we need so that the proportions work out. We want the
            #subplots to vary in width but the bars to be the same size. (It looks 
            #much nicer.) To do this we need the total number of x-axis ticks to give to 
            #the GridSpec constructor. 
            max_timelines = []
            for i, day in enumerate(self.days_dict.values()):
                df_day = data_term[data_term['Meetings'].apply(lambda x : day in x)].copy()
                df_day['start'] = df_day['Meetings'].apply(lambda x: x[day]['start'])
                df_day['end'] = df_day['Meetings'].apply(lambda x: x[day]['end'])
                df_day = df_day.sort_values('start', ascending=True)
                timelines = []
                for _, row in df_day.iterrows():
                    n = 1
                    while any ((s <= row['start'] <= e) and (idx == n) 
                               for (s, e, idx) in 
                               zip(df_day['start'], df_day['end'], timelines)):
                        n += 1
                    timelines.append(n)
                df_day['timeline'] = timelines
                max_timeline = df_day['timeline'].max()
                max_timelines.append(max_timeline)

            plot_start_time = data_term['Meetings'].apply(lambda d: min(d[k]['start'] for k in d)).min()
            plot_end_time = data_term['Meetings'].apply(lambda d: min(d[k]['end'] for k in d)).max()
            plot_height = (plot_end_time - plot_start_time).total_seconds() / 3600
            plot_width = (sum(np.nan_to_num(max_timelines, nan=1))*1.25)
            gs = gridspec.GridSpec(1, int(sum(np.nan_to_num(max_timelines, nan=1))))

            fig = plt.figure(figsize=(plot_width,plot_height))
            fig.suptitle(term)

            start_idx = 0

            for i, day in enumerate(self.days_dict.values()):
                df_day = data_term[data_term['Meetings'].apply(lambda x : day in x)].copy()
                df_day['start'] = df_day['Meetings'].apply(lambda x: x[day]['start'])
                df_day['end'] = df_day['Meetings'].apply(lambda x: x[day]['end'])
                df_day = df_day.sort_values('start', ascending=True)
                
                timelines = []
                for _, row in df_day.iterrows():
                    n = 1
                    while any ((s <= row['start'] <= e) and (idx == n) 
                               for (s, e, idx) in 
                               zip(df_day['start'], df_day['end'], timelines)):
                        n += 1
                    timelines.append(n)
                df_day['timeline'] = timelines

                max_timeline = df_day['timeline'].max()

                
                # Create a subplot with a width proportional to the maximum 'timeline' value
                try: ax = plt.subplot(gs[0, start_idx:start_idx + max_timeline])
                except TypeError: ax = plt.subplot(gs[0, start_idx:start_idx + 1])
                ax.set_title(self.days_dict_reverse[day])
                if start_idx > 0:
                    plt.setp(ax.get_yticklabels(), visible=False)
                for _, row in df_day.iterrows():
                    bottom = mdates.date2num(row['start'])
                    height = mdates.date2num(row['end']) - bottom
                    left = row['timeline']
                    bar = ax.bar(left, height, width=1, bottom=bottom, color=row['color'])


                    # if int(sum(np.nan_to_num(max_timelines, nan=1))) > 20:
                    #     fntsize = 4
                    # elif int(sum(np.nan_to_num(max_timelines, nan=1))) > 15:
                    #     fntsize = 6
                    # elif int(sum(np.nan_to_num(max_timelines, nan=1))) > 10:
                    #     fntsize=8
                    # elif int(sum(np.nan_to_num(max_timelines, nan=1))) > 4:
                    #     fntsize=10
                    
 
                    ax.text(left, bottom + height/2, 
                            #String manipulation to remove any crosslisted courses---
                            #easier than passing the original 'Course' value to the writer
                            re.sub(r" \(.*?\)", "", row['Course'])+'\n'+
                            row['surname']+'\n'+
                            str(row['start'].strftime("%I:%M"))+
                            '-'+str(row['end'].strftime("%I:%M")), 
                            ha='center', 
                            va='center', 
                            fontsize=7)
                ax.yaxis_date()
                ax.yaxis.set_major_formatter(mdates.DateFormatter('%I:%M%p'))
                ax.set_ylim([plot_start_time,plot_end_time])
                try: ax.set_xlim(df_day['timeline'].min()-0.5,df_day['timeline'].max()+0.5)
                except ValueError: pass
                ax.set_xlabel('')
                ax.set_xticks([])
                ax.set_ylabel('')
                ax.invert_yaxis() 

                start_idx += max_timeline
            plt.tight_layout()
            plt.subplots_adjust(wspace=0)

            # Save the plot as an image with a unique name
            
            img_filename = 'output{}.png'.format(self.img_idx)

            plt.savefig(img_filename, dpi=120)

            # Load the image
            img = Image(img_filename)
            output_cell = 'A' + str(self.output_row)
            # cell = self.ws.cell(row=self.output_row, column=utils.column_index_from_string('A'))
            # cell.width = img.width
            # cell.height = img.height

            img.anchor=(output_cell)
            self.ws.add_image(img)

            self.output_row += 75
            self.img_idx += 1

    def write_course_schedule(self):
    #Function to write the Course Schedule page
    
        for term in self.complete_data['Term'].unique():

            #Add the Term row to the active WS
            self.ws.append([term])

            #Format the Term row
            self.ws.cell(self.ws.max_row, 
                             column=1).fill = PatternFill('solid', fgColor='808080')
            self.ws.cell(self.ws.max_row, 
                             column=1).font = Font(size=20,bold=True)
            self.ws.row_dimensions[self.ws.max_row].height = 30

            #Create the DF to go under Term
            mask = self.complete_data['Term'] == term
            new_df = self.data[mask]

            #Add the columns row to active WS
            self.ws.append(new_df.columns.tolist())

            #Format the columns row
            self.ws.row_dimensions[self.ws.max_row].height = 15
            for col in range(1, len(self.data.columns.tolist())+1):
                self.ws.cell(self.ws.max_row, column=col).fill = PatternFill('solid', fgColor='A9A9A9')
                self.ws.cell(self.ws.max_row, column=col).font = Font(italic=True)

            #Write rows of the new Term-based DF; format them in alternating colors
            count = 0
            for _, row in new_df.iterrows():
                self.ws.append(row.tolist())
                for col in range(1, len(self.data.columns.tolist())+1):
                    if count % 2 == 0:
                        self.ws.cell(self.ws.max_row, column=col).fill = PatternFill('solid', fgColor='E8E8E8')
                    else:
                        self.ws.cell(self.ws.max_row, column=col).fill = PatternFill('solid', fgColor='D3D3D3')
                    self.ws.row_dimensions[self.ws.max_row].height = 30
                count += 1

            #Separate terms with 2 blank lines
            self.ws.append([])
            self.ws.append([])

        #Post-processing
        ##Resize cells
        self.auto_resize(self.ws)
        ##Merge Term header cells
        self.merge_headers(self.ws,self.complete_data['Term'].unique(),len(self.data.columns.tolist()))


    def write_by_day(self):

        for term in self.complete_data['Term'].unique():
        #First iterate over terms

            #Add the term row
            self.ws.append([term])

            #Format the Term row
            self.ws.cell(self.ws.max_row, 
                         column=1).fill = PatternFill('solid', 
                                                      fgColor='808080')
            self.ws.cell(self.ws.max_row, 
                         column=1).font = Font(size=20,
                                               bold=True)
            self.ws.row_dimensions[self.ws.max_row].height = 30

            for day in list(self.days_dict.keys()):
                #Then interate over days of the week

                #Add the day row
                self.ws.append([day])

                #Format the day row
                self.ws.cell(self.ws.max_row, 
                             column=1).fill = PatternFill('solid', 
                                                          fgColor='A9A9A9')
                self.ws.cell(self.ws.max_row, 
                             column=1).font = Font(bold=True)

                #Create a mask for by-day sub-DFs: Only retain rows in that term on that day
                mask = (self.complete_data['Term'] == term) & (self.complete_data['Meetings']
                                                               .apply(lambda x: self.days_dict[day] in x))   


                #It takes a few steps to sort the sub-DF by start time...
                ##First concatenate the 'meetings' column in order to retrieve the start time
                ##(Also apply the mask to this new df)
                ##(Concatenation SHOULDN'T be a problem because we haven't altered any columns yet.)
                new_df = pd.concat([self.data, 
                                     self.complete_data['Meetings']],
                                     axis=1)[mask]

                ##Next add a 'start' column, containing the start times of each meeting
                ##(This is a bit ugly because if you refer to [day] within the scope of 
                ##lambda x, then it doesn't retain its value at this point in the
                ##method because it's evaluated later---so we have to raise it (QR!) and
                ##plug it in over the lambda x.)
                new_df['start'] = new_df['Meetings'].apply((lambda d: 
                                                            (lambda x: 
                                                             x[d]['start']))
                                                             (self.days_dict[day]))
                
                ##Sort the rows...
                new_df.sort_values(by='start', inplace=True)

                ##And drop the new columns!
                new_df.drop(axis=1,columns=['Meetings','start'],inplace=True)

                #Add the columns row
                self.ws.append(new_df.columns.tolist())

                #Format the columns row
                for col in range(1, len(self.data.columns.tolist())+1):
                    self.ws.cell(self.ws.max_row, 
                                 column=col).fill = PatternFill('solid', 
                                                                fgColor='A9A9A9')
                    self.ws.cell(self.ws.max_row, column=col).font = Font(italic=True)


                #Append the DF row by row and fill with alternating colors
                count = 0
                for _, row in new_df.iterrows():
                    self.ws.append(row.tolist())
                    for col in range(1, len(self.data.columns.tolist())+1):
                        if count % 2 == 0:
                            self.ws.cell(self.ws.max_row, 
                                         column=col).fill = PatternFill('solid', 
                                                                        fgColor='E8E8E8')
                        else:
                            self.ws.cell(self.ws.max_row, 
                                         column=col).fill = PatternFill('solid', 
                                                                        fgColor='D3D3D3')
                    self.ws.row_dimensions[self.ws.max_row].height = 30
                    count += 1
            ##Terms are separated by two blank lines
            self.ws.append([])
            self.ws.append([])
        
        #Post-processing: resize cells, merge headers to extend across the entire chart
        self.auto_resize(self.ws)
        self.merge_headers(self.ws,self.complete_data['Term'].unique(),len(self.data.columns.tolist()))
        self.merge_headers(self.ws,list(self.days_dict.keys()),len(self.data.columns.tolist()))


    def write_by_instructor(self):

        for term in self.complete_data['Term'].unique():
        #First iterate over terms

            #Add the term row
            self.ws.append([term])

            #Format the Term row
            self.ws.cell(self.ws.max_row, 
                         column=1).fill = PatternFill('solid', 
                                                      fgColor='808080')
            self.ws.cell(self.ws.max_row, 
                         column=1).font = Font(size=20,
                                               bold=True)
            self.ws.row_dimensions[self.ws.max_row].height = 30

            
            for name in self.names:
            #Next iterate over instructor names

                #Mask for sub-DFs: same instructor and same term
                mask = ((self.complete_data['Term'] == term) &
                         (self.complete_data['Instructor'] == name))
                new_df = self.data[mask].copy()

                #Drop instructor ('term' never made it over to the writer)
                new_df.drop(['Instructor'], 
                            axis=1,
                            inplace=True
                            )

                if not new_df.empty:
                #Only write an instructor in a particular term if they have courses
                #in that term

                    #Add the instructor row
                    self.ws.append([name])

                    #Format the instructor row
                    self.ws.cell(self.ws.max_row, 
                                 column=1).fill = PatternFill('solid', 
                                                              fgColor='A9A9A9')
                    self.ws.cell(self.ws.max_row, 
                                 column=1).font = Font(bold=True)
                    
                    #Add the columns row
                    self.ws.append(new_df.columns.tolist())

                    #Format the columns row
                    for col in range(1, len(self.data.columns.tolist())):
                        self.ws.cell(self.ws.max_row, 
                                     column=col).fill = PatternFill('solid', 
                                                                    fgColor='A9A9A9')
                        self.ws.cell(self.ws.max_row, 
                                     column=col).font = Font(italic=True)


                    #Fill in the sub-DF and format with alternating colors
                    count = 0
                    for _, row in new_df.iterrows():
                        self.ws.append(row.tolist())
                        for col in range(1, len(self.data.columns.tolist())):
                            if count % 2 == 0:
                                self.ws.cell(self.ws.max_row, 
                                             column=col).fill = PatternFill('solid', 
                                                                            fgColor='E8E8E8')
                            else:
                                self.ws.cell(self.ws.max_row, 
                                             column=col).fill = PatternFill('solid', 
                                                                            fgColor='D3D3D3')
                        self.ws.row_dimensions[self.ws.max_row].height = 30
                        count += 1
            
            #Separate terms with two blank lines
            self.ws.append([])
            self.ws.append([])

        #Post-processing
        self.auto_resize(self.ws)
        self.merge_headers(self.ws,
                           self.complete_data['Term'].unique(),
                           len(self.data.columns.tolist())-1
                           )
        self.merge_headers(self.ws,
                           self.complete_data['Instructor'].unique(),
                           len(self.data.columns.tolist())-1
                           )
